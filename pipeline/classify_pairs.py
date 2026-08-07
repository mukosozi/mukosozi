#!/usr/bin/env python3
"""
Mukosozi correction-pair classifier (triage stage).

What it does, when the collected Excel comes in:
  1. Reads the "Corrections" sheet, or every daily tab named Corrections_* .
  2. Normalises whitespace, drops exact duplicates, groups near-duplicates,
     and flags conflicts (same wrong form mapped to two different fixes).
  3. Separates cosmetic normalization (final punctuation, sentence case) from
     the core change, then diffs each pair by SURFACE features only.
  4. Proposes a bucket (grammar / word choice / spelling / spacing /
     punctuation / capitalization) with a plain reason and a confidence.
  5. Writes a triage workbook, sorted by proposed bucket, with a blank
     "Your bucket" column for Christophe to make the real call.

It deliberately does NOT assert any Kinyarwanda fact. A one-letter change can be
a spelling slip or a grammar-agreement rule; the tool says so and leaves the
decision to the native speaker.

Usage:
    python classify_pairs.py collected.xlsx            # -> collected_triage.xlsx
    python classify_pairs.py collected.xlsx out.xlsx
"""

import sys, string, re
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

PUNCT = set(string.punctuation) | {"’", "‘", "“", "”", "«", "»", "–", "—"}

# ---------- string helpers ----------
def norm_ws(s):
    return re.sub(r"\s+", " ", str(s if s is not None else "").strip())

def strip_punct(s):
    return "".join(ch for ch in s if ch not in PUNCT)

def despace(s):
    return re.sub(r"\s+", "", s)

def edit_distance(a, b):
    if a == b: return 0
    la, lb = len(a), len(b)
    if la == 0: return lb
    if lb == 0: return la
    prev = list(range(lb + 1))
    for i in range(1, la + 1):
        cur = [i] + [0] * lb
        for j in range(1, lb + 1):
            cost = 0 if a[i-1] == b[j-1] else 1
            cur[j] = min(prev[j] + 1, cur[j-1] + 1, prev[j-1] + cost)
        prev = cur
    return prev[lb]

# ---------- the diff / classify core ----------
def analyze(incorrect, correct, unit=""):
    inc, cor = norm_ws(incorrect), norm_ws(correct)
    unit = (unit or "").strip().lower()
    out = {"edit_type": "", "changed_from": "", "changed_to": "",
           "token_change": "", "edit_distance": edit_distance(inc, cor),
           "normalized": "", "bucket": "", "section_hint": "", "confidence": "", "why": ""}

    if inc == cor:
        out.update(edit_type="no change", bucket="skip", confidence="high",
                   section_hint="none",
                   why="Incorrect and correct are identical after trimming. Check this row.")
        return out

    # --- isolate cosmetic normalization (final punctuation, sentence case) from the core change ---
    TERM = " .!?…,;:"
    inc2, cor2 = inc.rstrip(TERM), cor.rstrip(TERM)
    flags = []
    if inc[len(inc2):] != cor[len(cor2):]:
        flags.append("final punctuation")
    if inc2 and cor2 and ((inc2.lower() == cor2.lower() and inc2 != cor2)
                          or (inc2[0].islower() and cor2[0].isupper())):
        flags.append("capitalization")
    out["normalized"] = " + ".join(flags)
    tail = (" The fix also normalizes " + " and ".join(flags) + ".") if flags else ""
    li, lc = inc2.lower(), cor2.lower()

    # cosmetic only: core text identical
    if li == lc:
        b = "capitalization" if "capitalization" in flags else "punctuation"
        out.update(edit_type="cosmetic only", bucket=b, confidence="high",
                   section_hint="generic normalization (or lexical exception if a specific rule)",
                   why="Core text is identical; only " + (" and ".join(flags) or "case/punctuation") + " changed.")
        return out

    # internal punctuation only (same letters, same word boundaries)
    if strip_punct(li) == strip_punct(lc):
        out.update(edit_type="punctuation (internal)", bucket="punctuation", confidence="medium",
                   section_hint="lexical exception (contraction/orthography) or normalization",
                   why="Same letters and spacing; internal punctuation differs." + tail)
        return out

    # spacing (merge/split)
    if despace(li) == despace(lc):
        out.update(edit_type="spacing (merge/split)", bucket="spacing", confidence="high",
                   section_hint="lexical exception (spacing) or generic normalization",
                   why="Same letters, only word boundaries differ (words merged or split)." + tail)
        return out
    if despace(strip_punct(li)) == despace(strip_punct(lc)):
        out.update(edit_type="spacing + punctuation", bucket="spacing", confidence="medium",
                   section_hint="lexical exception (spacing/contraction)",
                   why="Word boundaries and internal punctuation differ; the letters are the same." + tail)
        return out

    it, ct = li.split(), lc.split()
    # same number of tokens -> look at which changed
    if len(it) == len(ct):
        diffs = [(a, b) for a, b in zip(it, ct) if a != b]
        if len(diffs) == 1:
            a, b = diffs[0]
            d = edit_distance(a, b)
            out.update(changed_from=a, changed_to=b, token_change="1 of %d tokens" % len(it))
            if d <= 2:
                out.update(edit_type="single-token spelling", bucket="spelling",
                           confidence="medium",
                           section_hint="lexical exception (misspelling) or spelling rule",
                           why=("One token changed by a small edit (distance %d)." % d) + tail)
            else:
                out.update(edit_type="single-token word change", bucket="word choice",
                           confidence="low",
                           section_hint="lexical exception or grammar rule",
                           why=("One token replaced by a fairly different one (distance %d)." % d) + tail)
            # a single short-token change inside a sentence is often grammar, not spelling
            if unit == "sentence" or len(it) >= 4:
                out["confidence"] = "low"
                out["why"] += " Sits inside a sentence, so this may be a grammar/agreement rule rather than a spelling fix."
            return out
        else:
            froms = " | ".join(a for a, _ in diffs)
            tos = " | ".join(b for _, b in diffs)
            out.update(edit_type="multi-token change", bucket="grammar",
                       changed_from=froms, changed_to=tos,
                       token_change="%d of %d tokens" % (len(diffs), len(it)),
                       confidence="low",
                       section_hint="grammar rule (agreement/structure); could be lexical if one root drives it",
                       why=("%d tokens changed together, which often means an agreement or structural rule." % len(diffs)) + tail)
            return out

    # token counts differ
    dd = edit_distance(despace(li), despace(lc))
    if dd <= 3:
        out.update(edit_type="split/merge with minor change", bucket="spacing",
                   token_change="%d -> %d tokens" % (len(it), len(ct)), confidence="low",
                   section_hint="lexical exception (spacing) or grammar; review",
                   why=("Word count changed with only a small letter change (distance %d)." % dd) + tail)
    else:
        out.update(edit_type="insertion/deletion / multi-token", bucket="grammar",
                   token_change="%d -> %d tokens" % (len(it), len(ct)), confidence="low",
                   section_hint="grammar rule (structure); review",
                   why="Word count changed with larger edits, likely a structural or grammar rule." + tail)
    return out

# ---------- loading the collection ----------
def _find_header(ws):
    """A real header row has 'incorrect' and 'correct' in two DIFFERENT cells.
    (The row-1 banner mentions both words in one merged cell, so it must not match.)"""
    for r in range(1, 7):
        cols = {}
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(r, c).value or "").strip().lower()
            if not v: continue
            if "incorrect" in v and "incorrect" not in cols: cols["incorrect"] = c
            elif "correct" in v and "correct" not in cols: cols["correct"] = c
            elif "source" in v and "source" not in cols: cols["source"] = c
            elif v.startswith("unit"): cols["unit"] = c
            elif "context" in v: cols["context"] = c
            elif v == "id": cols["id"] = c
        if "incorrect" in cols and "correct" in cols and cols["incorrect"] != cols["correct"]:
            return r, cols
    return None, None

FALLBACK_COLS = {"id": 1, "incorrect": 2, "correct": 3, "source": 4, "unit": 5, "context": 6}
HEADER_WORDS = {"incorrect (as written)", "correct (fixed)"}

def load_collection(path):
    """Reads one 'Corrections' sheet, or every sheet named Corrections_* (daily tabs)."""
    wb = load_workbook(path, data_only=True)
    sheets = [s for s in wb.sheetnames if s.lower().startswith("corrections")] or [wb.sheetnames[0]]
    rows = []
    for name in sheets:
        ws = wb[name]
        header_row, cols = _find_header(ws)
        if header_row is None:
            # headers were edited/cleared; fall back to the template's known layout
            header_row, cols = 2, FALLBACK_COLS
        tag = name.replace("Corrections", "").strip("_ ")
        for r in range(header_row + 1, ws.max_row + 1):
            inc = ws.cell(r, cols["incorrect"]).value
            cor = ws.cell(r, cols["correct"]).value
            if not norm_ws(inc) or not norm_ws(cor):
                continue
            if norm_ws(inc).lower() in HEADER_WORDS or norm_ws(cor).lower() in HEADER_WORDS:
                continue  # structural remnant, not data
            rid = ws.cell(r, cols["id"]).value if "id" in cols else r
            rows.append({
                "id": f"{tag}:{rid}" if tag else rid,
                "incorrect": inc, "correct": cor,
                "source": ws.cell(r, cols["source"]).value if "source" in cols else "",
                "unit": ws.cell(r, cols["unit"]).value if "unit" in cols else "",
                "context": ws.cell(r, cols["context"]).value if "context" in cols else "",
            })
    if not rows:
        raise SystemExit("No data rows found (looked for sheets starting with 'Corrections').")
    return rows

# ---------- dedup + conflicts ----------
def dedup(rows):
    """Groups on (incorrect, CORE of correct): per ruling (Aug 2026), terminal
    punctuation is the writer's tone, so fixes differing only in . ! ... or
    final case merge into one entry with their variants recorded."""
    TERMC = " .!?\u2026,;:"
    groups = defaultdict(list)
    by_inc = defaultdict(set)
    for row in rows:
        ni = norm_ws(row["incorrect"]).lower()
        core = norm_ws(row["correct"]).lower().rstrip(TERMC)
        groups[(ni, core)].append(row)
        by_inc[ni].add(core)
    unique = []
    for (ni, core), members in groups.items():
        first = members[0]
        variants = sorted({norm_ws(m["correct"]) for m in members})
        conflict = len(by_inc[ni]) > 1
        unique.append({**first, "count": len(members),
                       "source_ids": ", ".join(str(m["id"]) for m in members),
                       "variants": variants, "conflict": conflict})
    return unique

# ---------- write triage workbook ----------
GOLD="D4A853"; DARK="0D0D0D"; BANNER="F6EBCB"; WARN="F6C6C6"
def F(**k): return Font(name="Arial", **k)
BUCKET_ORDER = {"grammar":0,"word choice":1,"spelling":2,"spacing":3,
                "punctuation":4,"capitalization":5,"skip":6}
YOUR_BUCKET='"grammar rule,lexical exception,neologism,generic-normalization,duplicate,skip"'

def write_triage(analyzed, out_path):
    wb = Workbook()
    thin = Side(style="thin", color="D9D9D9"); border = Border(thin, thin, thin, thin)
    # Summary sheet
    summ = wb.active; summ.title = "Summary"; summ.sheet_view.showGridLines = False
    summ.column_dimensions["A"].width = 34; summ.column_dimensions["B"].width = 12
    summ.cell(1,1,"Mukosozi triage summary").font = F(bold=True, size=14, color="C75D3A")
    counts = defaultdict(int); conflicts = 0
    for a in analyzed:
        counts[a["bucket"]] += 1
        if a.get("conflict"): conflicts += 1
    summ.cell(3,1,"Unique pairs").font=F(bold=True); summ.cell(3,2,len(analyzed))
    summ.cell(4,1,"Conflicts (same wrong form, different fix)").font=F(bold=True,color="B00020"); summ.cell(4,2,conflicts)
    summ.cell(6,1,"Proposed bucket").font=F(bold=True); summ.cell(6,2,"Count").font=F(bold=True)
    r=7
    for b in sorted(counts, key=lambda x: BUCKET_ORDER.get(x,9)):
        summ.cell(r,1,b); summ.cell(r,2,counts[b]); r+=1

    # Review sheet
    ws = wb.create_sheet("Review")
    headers = ["Src IDs","Count","Conflict?","Incorrect","Correct","Source","Unit",
               "Proposed bucket","Edit type","Also normalized","Changed from","Changed to","Tokens",
               "Edit dist","Confidence","Likely section","Why","Your bucket","Your notes"]
    widths = [10,7,9,34,34,14,10,16,22,15,18,18,12,9,11,34,44,18,30]
    for i,h in enumerate(headers,1):
        c=ws.cell(2,i,h); c.font=F(bold=True,color=DARK,size=10); c.fill=PatternFill("solid",fgColor=GOLD)
        c.alignment=Alignment(wrap_text=True,vertical="center"); c.border=border
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    b=ws.cell(1,1,"Auto-sorted triage. 'Proposed bucket' is a surface-feature guess, not a linguistic verdict. "
                  "Set 'Your bucket' to make the real call. Red rows = the same wrong form was corrected two different ways, decide which is right.")
    b.font=F(bold=True,color="6B4E1E",size=10); b.fill=PatternFill("solid",fgColor=BANNER)
    b.alignment=Alignment(wrap_text=True,vertical="center"); ws.row_dimensions[1].height=42
    analyzed_sorted = sorted(analyzed, key=lambda a:(BUCKET_ORDER.get(a["bucket"],9), str(a["incorrect"]).lower()))
    row=3
    for a in analyzed_sorted:
        why = a["why"]
        if len(a.get("variants", [])) > 1:
            why += "  Punctuation variants merged per ruling (tone): " + " | ".join(a["variants"])
        vals=[a["source_ids"],a["count"],"YES" if a["conflict"] else "",
              norm_ws(a["incorrect"]),norm_ws(a["correct"]),a.get("source") or "",a.get("unit") or "",
              a["bucket"],a["edit_type"],a.get("normalized",""),a["changed_from"],a["changed_to"],a["token_change"],
              a["edit_distance"],a["confidence"],a["section_hint"],why,"",""]
        for i,v in enumerate(vals,1):
            c=ws.cell(row,i,v); c.font=F(size=10); c.border=border
            c.alignment=Alignment(wrap_text=True,vertical="top",
                                  horizontal="center" if i in (2,3,13,14,15) else "left")
        if a["conflict"]:
            for i in range(1,len(headers)+1):
                ws.cell(row,i).fill=PatternFill("solid",fgColor=WARN)
        row+=1
    dv=DataValidation(type="list",formula1=YOUR_BUCKET,allow_blank=True)
    ws.add_data_validation(dv); _yb=get_column_letter(len(headers)-1); dv.add(f"{_yb}3:{_yb}{row-1}")
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes="D3"; ws.auto_filter.ref=f"A2:{get_column_letter(len(headers))}{row-1}"
    wb.save(out_path)

def main():
    if len(sys.argv) < 2:
        print(__doc__); sys.exit(1)
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else src.rsplit(".",1)[0] + "_triage.xlsx"
    rows = load_collection(src)
    unique = dedup(rows)
    for u in unique:
        u.update(analyze(u["incorrect"], u["correct"], u.get("unit")))
    write_triage(unique, out)
    dupes = len(rows) - len(unique)
    print(f"read {len(rows)} pairs -> {len(unique)} unique ({dupes} duplicates removed)")
    print(f"wrote {out}")

if __name__ == "__main__":
    main()
