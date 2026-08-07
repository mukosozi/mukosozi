#!/usr/bin/env python3
"""
Mukosozi word-bank exporter, v1.1.

Turns Christophe's reviewed triage rows into word_bank.json entries with their
evidence attached in attestations.json.

Reads:
  1. the triage workbook  (Review sheet; only rows where "Your bucket" is filled)
  2. the collection workbook (daily tabs; source of every attestation)
  3. the current word_bank.json (hand-made entries are appended to, never touched)

Writes (to --out dir, default ./out):
  word_bank.json        merged bank, schema-identical to the current file
  attestations.json     evidence per entry id; receipts are never rewritten,
                        reviewer rulings are bridged via notes
  export_report.txt     ids, emendations, flips, flags, skips

v1.1 adds:
  - STICKY IDS: previously assigned ids are reused via the index of the prior
    out/attestations.json, so shipped ids never renumber
  - REVIEWER EMENDATION: when the ruled correct form differs from what the
    contributor wrote, the entry carries the ruled form and each receipt keeps
    the attested text plus a note naming the ruling
  - FLIP ("flip" in Your notes): the observed wild form is standard and the
    contributor's correction was the error; the entry protects the wild form
    (e.g. compounds that must not be split)

Policies encoded (per Christophe's rulings, Aug 2026): core-form headwords,
terminal punctuation is the writer's tone, sentence capitalization is the
normalization layer's job, only Christophe's verdicts export, reruns are safe.

Usage:
    python export_word_bank.py triage.xlsx collection.xlsx word_bank.json [--out DIR]
"""

import sys, os, json, argparse, datetime
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from classify_pairs import load_collection, norm_ws

TERMC = " .!?\u2026,;:"
TODAY = datetime.date.today().isoformat()
REVIEWER = "Christophe Mumaragishyika"
COLLECTOR = "Sarah Izabayo"

VERDICTS = {"lexical exception", "grammar rule", "neologism",
            "generic-normalization", "duplicate", "skip"}


def core(s):
    return norm_ws(s).lower().rstrip(TERMC)


def core_form(incorrect, correct):
    """Strip terminal punctuation; lower any leading capital (positional case).
    Every lowering is flagged so proper nouns get a human glance."""
    c = norm_ws(correct).rstrip(TERMC)
    flagged = False
    if c and c[0].isupper():
        c = c[0].lower() + c[1:]
        flagged = True
    return c, flagged


def read_reviewed(triage_path):
    wb = load_workbook(triage_path, data_only=True)
    ws = wb["Review"]
    rows = []
    for r in range(3, ws.max_row + 1):
        verdict = norm_ws(ws.cell(r, 18).value).lower()
        if not verdict:
            continue
        rows.append({
            "row": r,
            "src_ids": norm_ws(ws.cell(r, 1).value),
            "incorrect": norm_ws(ws.cell(r, 4).value),
            "correct": norm_ws(ws.cell(r, 5).value),
            "source": norm_ws(ws.cell(r, 6).value),
            "unit": norm_ws(ws.cell(r, 7).value).lower(),
            "proposed": norm_ws(ws.cell(r, 8).value).lower(),
            "edit_type": norm_ws(ws.cell(r, 9).value).lower(),
            "verdict": verdict,
            "notes": norm_ws(ws.cell(r, 19).value),
        })
    return rows


def build_attestations(src_ids, coll_index, ruled_core):
    """One receipt per raw collection row. Receipts keep the attested text;
    if the reviewer's ruled form differs, a note bridges them."""
    atts, missing = [], []
    for sid in [s.strip() for s in src_ids.split(",") if s.strip()]:
        raw = coll_index.get(sid)
        if raw is None:
            missing.append(sid)
            continue
        day = sid.split(":")[0] if ":" in sid else ""
        att = {
            "observed": norm_ws(raw["incorrect"]),
            "text": norm_ws(raw["correct"]),
            "gloss": norm_ws(raw.get("context")),
            "source": norm_ws(raw.get("source")),
            "date": f"2026-{day.replace('_', '-')}" if day else "",
        }
        if ruled_core and core(att["text"]) != ruled_core:
            att["note"] = f"reviewer ruling: standard form is '{ruled_core}'"
        atts.append(att)
    return atts, missing


def flip_attestations(src_ids, coll_index):
    """For flipped rows the wild observation IS the standard form."""
    atts, missing = [], []
    for sid in [s.strip() for s in src_ids.split(",") if s.strip()]:
        raw = coll_index.get(sid)
        if raw is None:
            missing.append(sid)
            continue
        day = sid.split(":")[0] if ":" in sid else ""
        atts.append({
            "observed": norm_ws(raw["incorrect"]),
            "text": norm_ws(raw["incorrect"]),
            "gloss": norm_ws(raw.get("context")),
            "source": norm_ws(raw.get("source")),
            "date": f"2026-{day.replace('_', '-')}" if day else "",
            "note": ("attested in the standard form; the split in the "
                     "collection file was a collection-side error"),
        })
    return atts, missing


def lexical_entry(eid, ruled_word, row, atts, flipped=False, emended_from=None, gloss_override=None):
    if flipped:
        rule = (f"Standard form: '{ruled_word}' (one word). "
                f"The split form '{norm_ws(row['correct']).rstrip(TERMC)}' is invalid.")
        incorrect_ex = [{"text": row["correct"],
                         "error": "Invalid split; this compound is one word"}]
        correct_ex = [{"text": a["text"], "translation": a["gloss"]} for a in atts[:3]]
        category = "spacing_rules"
    else:
        spacing = "spacing" in row["edit_type"] or row["proposed"] == "spacing"
        category = "spacing_rules" if spacing else (
            "spelling" if row["proposed"] == "spelling" else "word_choice")
        if spacing:
            rule = (f"Standard form: '{ruled_word}'. The form '{row['incorrect']}' "
                    f"joins or splits word boundaries incorrectly.")
            err = "Word-boundary error"
        else:
            rule = (f"Standard form: '{ruled_word}'. The form '{row['incorrect']}' "
                    f"is a non-standard spelling.")
            err = "Non-standard form"
        if emended_from:
            rule += (f" Reviewer emendation: the collected correction "
                     f"'{emended_from}' was amended to '{ruled_word}'.")
        incorrect_ex = [{"text": row["incorrect"], "error": err}]
        correct_ex = [{"text": row["correct"],
                       "translation": gloss_override or (atts[0]["gloss"] if atts else "")}]
        for a in atts:
            if len(correct_ex) >= 3:
                break
            if core(a["text"]) == core(ruled_word) and a["text"] != row["correct"]:
                correct_ex.append({"text": a["text"], "translation": a["gloss"]})
    entry = {
        "id": eid,
        "word": ruled_word,
        "category": category,
        "definition": gloss_override or (atts[0]["gloss"] if atts else ""),
        "rule": rule,
        "examples": {"correct": correct_ex, "incorrect": incorrect_ex},
        "contributor": COLLECTOR,
        "reviewed_by": REVIEWER,
        "status": "verified",
        "added_date": TODAY,
    }
    if row["unit"] == "phrase":
        entry["pos"] = "phrase"
    if row["notes"]:
        entry["notes"] = row["notes"]
    return entry


def grammar_entry(eid, ruled_word, row, atts):
    return {
        "id": eid,
        "name": f"Imported from corpus: {ruled_word}",
        "category": "corpus_imported",
        "description": atts[0]["gloss"] if atts else "",
        "rule": row["notes"] or "(rule write-up pending)",
        "examples": {
            "correct": [{"text": a["text"], "translation": a["gloss"]}
                        for a in atts[:3]],
            "incorrect": [{"text": row["incorrect"],
                           "error": "See corrected form"}],
        },
        "contributor": COLLECTOR,
        "reviewed_by": REVIEWER,
        "status": "under_review",
        "added_date": TODAY,
    }


def neologism_entry(eid, ruled_word, row, atts):
    return {
        "id": eid,
        "proposed_word": ruled_word,
        "definition": atts[0]["gloss"] if atts else "",
        "category": "corpus_imported",
        "morphological_analysis": "(pending)",
        "rationale": row["notes"] or "(pending)",
        "alternatives_considered": [],
        "status": "proposed",
        "notes": "Imported from the July 2026 correction corpus.",
        "contributor": COLLECTOR,
        "reviewed_by": REVIEWER,
        "added_date": TODAY,
    }


def next_counters(bank, ledger):
    counters = {}
    ids = [e.get("id", "") for section in
           ("grammar_rules", "lexical_exceptions", "neologisms")
           for e in bank.get(section, [])] + list(ledger.values())
    for eid in ids:
        if "-" in eid:
            p, n = eid.rsplit("-", 1)
            if n.isdigit():
                counters[p] = max(counters.get(p, 0), int(n))
    return counters


def known_forms(bank):
    forms = set()
    for section in ("grammar_rules", "lexical_exceptions", "neologisms"):
        for e in bank.get(section, []):
            for ex in e.get("examples", {}).get("incorrect", []):
                forms.add(norm_ws(ex.get("text")).lower())
    return forms


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("triage")
    ap.add_argument("collection")
    ap.add_argument("word_bank")
    ap.add_argument("--out", default="out")
    args = ap.parse_args()
    os.makedirs(args.out, exist_ok=True)

    bank = json.load(open(args.word_bank, encoding="utf-8"))
    att_path = os.path.join(args.out, "attestations.json")

    # sticky ids: seed the ledger from the previous export's index
    ledger = {}
    if os.path.exists(att_path):
        try:
            ledger = dict(json.load(open(att_path, encoding="utf-8")).get("index", {}))
        except Exception:
            ledger = {}

    attdoc = {
        "metadata": {
            "name": "Mukosozi Attestation Bank",
            "description": ("Evidence behind word-bank entries: each item is a "
                            "dated, sourced observation with the contributor's "
                            "English gloss. Receipts keep the text exactly as "
                            "attested; reviewer rulings are bridged via notes. "
                            "Terminal punctuation is preserved (writer's tone)."),
            "license": bank["metadata"].get("license", "CC BY-NC-SA 4.0"),
            "collection": "July 2026 correction corpus",
            "collector": COLLECTOR,
            "reviewer": REVIEWER,
            "repository": bank["metadata"].get("repository", ""),
            "created": TODAY,
        },
        "index": {},
        "entries": {},
    }

    coll_index = {str(r["id"]): r for r in load_collection(args.collection)}
    reviewed = read_reviewed(args.triage)
    counters = next_counters(bank, ledger)
    existing = known_forms(bank)

    stats = {"exported": [], "excluded": 0, "already": 0, "unknown": [],
             "flags": [], "missing": [], "grammar_pending": [],
             "reused_ids": 0, "emended": [], "flipped": []}

    for row in reviewed:
        v = row["verdict"]
        if v not in VERDICTS:
            stats["unknown"].append((row["row"], v))
            continue
        if v in ("duplicate", "skip", "generic-normalization"):
            stats["excluded"] += 1
            continue
        key = row["incorrect"].lower()
        if key in existing:
            stats["already"] += 1
            continue
        if key in attdoc["index"]:
            stats["unknown"].append((row["row"], "same incorrect form verdicted twice"))
            continue

        flipped = "flip" in row["notes"].lower()

        gloss_override = None
        if "[gloss]" in row["notes"]:
            before, after = row["notes"].split("[gloss]", 1)
            gloss_override = after.strip()
            row["notes"] = before.strip()

        if flipped:
            ruled_word, flagged = core_form(row["correct"], row["incorrect"])
            atts, missing = flip_attestations(row["src_ids"], coll_index)
        else:
            ruled_word, flagged = core_form(row["incorrect"], row["correct"])
            atts, missing = build_attestations(row["src_ids"], coll_index, ruled_word)
        stats["missing"] += missing

        # reviewer emendation: ruled form differs from every attested correction
        emended_from = None
        if not flipped and atts and all(core(a["text"]) != ruled_word for a in atts):
            emended_from = norm_ws(atts[0]["text"]).rstrip(TERMC)
            stats["emended"].append((row["row"], emended_from, ruled_word))

        if v == "lexical exception":
            spacing = flipped or "spacing" in row["edit_type"] or row["proposed"] == "spacing"
            prefix = "SP" if spacing else "LE"
        elif v == "grammar rule":
            prefix = "GR"
        else:
            prefix = "NEO"

        eid = ledger.get(key)
        if eid:
            stats["reused_ids"] += 1
        else:
            counters[prefix] = counters.get(prefix, 0) + 1
            eid = f"{prefix}-{counters[prefix]:03d}"
            ledger[key] = eid

        if v == "lexical exception":
            bank["lexical_exceptions"].append(
                lexical_entry(eid, ruled_word, row, atts, flipped, emended_from, gloss_override))
            if flipped:
                stats["flipped"].append(eid)
        elif v == "grammar rule":
            bank["grammar_rules"].append(grammar_entry(eid, ruled_word, row, atts))
            stats["grammar_pending"].append(eid)
        else:
            bank["neologisms"].append(neologism_entry(eid, ruled_word, row, atts))

        attdoc["entries"][eid] = atts
        attdoc["index"][key] = eid
        if flagged:
            stats["flags"].append((eid, ruled_word))
        stats["exported"].append(eid)

    bank["metadata"]["last_export"] = TODAY
    json.dump(bank, open(os.path.join(args.out, "word_bank.json"), "w",
              encoding="utf-8"), indent=2, ensure_ascii=False)
    json.dump(attdoc, open(att_path, "w", encoding="utf-8"),
              indent=2, ensure_ascii=False)

    lines = [
        f"Mukosozi export report — {TODAY}",
        f"reviewed rows found: {len(reviewed)}",
        f"exported: {len(stats['exported'])}  ({', '.join(stats['exported']) or '-'})",
        f"ids reused from ledger (stable): {stats['reused_ids']}",
        f"excluded (duplicate/skip/generic-normalization): {stats['excluded']}",
        f"already in the hand-made bank (skipped): {stats['already']}",
        f"unknown verdicts: {stats['unknown'] or '-'}",
        f"attestation ids not found in collection: {stats['missing'] or '-'}",
        "",
        "REVIEWER EMENDATIONS (entry carries the ruled form; receipts keep",
        "the attested text with a bridging note):",
    ] + [f"  row {r}: '{a}' -> '{b}'" for r, a, b in stats["emended"]] + [
        "",
        "FLIPPED ENTRIES (wild form is standard; split was the error):",
    ] + [f"  {eid}" for eid in stats["flipped"]] + [
        "",
        "PROPER-NOUN CHECK — first letter was lowered on these entries;",
        "restore the capital if any is a proper noun:",
    ] + [f"  {eid}: {w}" for eid, w in stats["flags"]] + [
        "",
        "GRAMMAR ENTRIES PENDING RULE TEXT (status under_review):",
    ] + [f"  {eid}" for eid in stats["grammar_pending"]]
    report = "\n".join(lines)
    open(os.path.join(args.out, "export_report.txt"), "w",
         encoding="utf-8").write(report)
    print(report)


if __name__ == "__main__":
    main()
